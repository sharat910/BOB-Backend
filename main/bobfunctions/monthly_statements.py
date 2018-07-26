from openpyxl import load_workbook
from django.conf import settings
from pprint import pprint
from datetime import date,datetime
import requests
import json
import xlrd
import calendar
from .fix_borders import patch_worksheet
import os

FILE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_ROOT = FILE_DIR + '/form_templates'
TEMPLATE_DIR = 'Monthly'

def fetch_batch():
    r = requests.get('http://localhost:8000/api/batch/2')
    batch_dict = json.loads(r.text)
    with open('batch.json', 'w') as outfile:
     json.dump(batch_dict, outfile, sort_keys = True, indent = 4,
               ensure_ascii = False)
    return batch_dict

def fetch_royaltyrate():
    r = requests.get('http://localhost:8000/api/royaltyrate/get_rate')
    rate = json.loads(r.text)['month_royalty']
    return rate


class MonthlyStatement(object):
    """docstring for MonthlyStatement."""
    def __init__(self, batch, month):
        super(MonthlyStatement, self).__init__()
        self.batch = batch
        self.month = month
        self.wb = None
        self.ws = None

    def generate_monthly_statement(self):
        self.set_wb_ws()
        self.fill_top()
        self.fill_students()
        file_path = self.get_final_file_path()
        self.wb.save(file_path)
        return self.generate_url(file_path)

    def generate_url(self,file_path):
        url = "/".join(file_path.split("/")[3:])
        return url

    def get_template_file_path(self,type,file_type='xlsx'):
        return "%s/%s/FormB_%s.%s" % (TEMPLATE_ROOT,TEMPLATE_DIR,type,file_type)

    def get_final_file_path(self,file_type='xlsx'):
        return settings.MEDIA_ROOT + 'monthly_statements/%s_%s_%s.%s' \
        % (self.batch['teacher_name'].replace(" ","_"),self.batch['level_detail'],\
        calendar.month_abbr[self.month],file_type)

    def bob_or_little_bob(self,level):
        if level <= 10:
            return 'bob'
        else:
            return 'little_bob'

    def get_excel_workbook(self,level):
        type = self.bob_or_little_bob(level)
        path = self.get_template_file_path(type)
        patch_worksheet()
        return load_workbook(path)

    def set_wb_ws(self):
        self.wb = self.get_excel_workbook(self.batch['level'])
        self.ws = self.wb.active

    def get_month(self,date_str,string=False):
        dt = datetime.strptime(date_str,"%Y-%m-%d")
        if string:
            return dt.strftime("%b")
        return dt.month

    def fill_top(self):
        today = date.today().strftime("%m/%d/%Y")
        self.ws['S5'] = today
        centre_name = str(self.batch['centre_exp']['name']).upper()
        centre_area = str(self.batch['centre_exp']['area']).upper()
        self.ws['J6'] = centre_name
        self.ws['Q6'] = centre_area
        self.ws['I7'] = self.batch['id']
        self.ws['N7'] = self.batch['level_detail']
        self.ws['P7'] = self.batch['day']
        self.ws['T7'] = self.batch['timing']
        self.ws['I8'] = str(self.batch['centre_exp']['code'])
        self.ws['M8'] = self.batch['teacher_name']
        self.ws['S8'] = datetime.strptime(self.batch['level_start_date'],"%Y-%m-%d").strftime("%m/%d/%Y")

    def fill_summary(self,total_data):
        royalty = fetch_royaltyrate()
        self.ws['L30'] = royalty
        self.ws['L31'] = royalty
        self.ws['L32'] = royalty

    def fetch_current_level_feerecords(self,student):
        return filter(lambda feerecord: feerecord['level'] == self.batch['level'],\
                    student['feerecords'])

    def get_student_row_data(self,batch_running_months,student):
        data = {
        'code': student['code'],
        'name': student['name']
        }

        feerecords = self.fetch_current_level_feerecords(student)
        if student['dropped']:
            if self.get_month(student['date_dropped']) == self.month:
                # Student dropped this month
                data['dropped'] = "DROPPED"
            else:
                return None
        else:
            data['dropped'] = ""

        payment_row = ['NP','NP','NP']
        for feerecord in feerecords:
            if feerecord['fee_type'] == 'Level':
                month_of_payment = self.get_month(feerecord['date_of_payment'])
                if month_of_payment == self.month:
                    payment_row = (3*[feerecord['fee_receipt_no']])
                elif month_of_payment < self.month:
                    payment_row = (3*['PAID'])
                else:
                    payment_row = (3*['NP'])
            elif feerecord['fee_type'] == 'Month':
                month_of_payment = self.get_month(feerecord['date_of_payment'])
                month_paid_for = feerecord['months'][0]
                assert month_paid_for in batch_running_months
                idx = batch_running_months.index(month_paid_for)
                if month_of_payment == self.month:
                    payment_row[idx] = feerecord['fee_receipt_no']
                elif month_of_payment < self.month:
                    payment_row[idx] = 'PAID'
                else:
                    payment_row[idx] = 'NP'
        assert len(payment_row) == 3
        data['payment'] = payment_row
        return data

    def fill_students(self):
        batch_running_months = self.batch['running_months']
        students = sorted(self.batch['students'],key= lambda x: x['code'])
        row_num = 13
        total_data = []
        for student in students:
            row_data = self.get_student_row_data(batch_running_months,student)
            print(row_data)
            if row_data:
                self.fill_row(row_data,row_num)
                total_data.append(row_data)
                row_num+=1
            else:
                pprint(student)
        self.fill_summary(total_data)

    def fill_row(self,row_data,row_num):
        self.ws['G%d' % row_num] = row_data['code']
        self.ws['I%d' % row_num] = row_data['name']
        self.ws['O%d' % row_num] = row_data['dropped']
        self.ws['Q%d' % row_num] = row_data['payment'][0]
        self.ws['R%d' % row_num] = row_data['payment'][1]
        self.ws['T%d' % row_num] = row_data['payment'][2]

if __name__ == '__main__':
    batch = fetch_batch()
    m = MonthlyStatement(batch,7)
    print(m.generate_monthly_statement())
